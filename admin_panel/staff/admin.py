import re

from django.contrib import admin
from django.db.models import Count
from django.http import HttpResponse
from django.utils import timezone
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from .models import AiImage, Image, Product, Staff, Store


def _strip_html(value: str) -> str:
    return re.sub(r'<[^>]+>', '', str(value))


def _img(url: str, size: int = 50) -> str:
    """Картинка с кликом — открывает оригинал в новой вкладке."""
    return (
        f'<a href="{url}" target="_blank" title="Открыть полный размер">'
        f'<img src="{url}" width="{size}" height="{size}" '
        f'style="object-fit:cover;border-radius:5px;margin:1px;'
        f'cursor:zoom-in;transition:opacity .15s;" '
        f'onmouseover="this.style.opacity=\'0.8\'" '
        f'onmouseout="this.style.opacity=\'1\'" />'
        f'</a>'
    )


def _safe_img(field, size: int = 50):
    if not field:
        return '—'
    try:
        return mark_safe(_img(field.url, size))
    except (ValueError, AttributeError):
        return 'Файл не найден'

def export_to_excel(modeladmin, request, queryset):
    opts = modeladmin.model._meta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename={opts.verbose_name_plural}.xlsx'
    )
    wb = Workbook()
    ws = wb.active
    ws.title = str(opts.verbose_name_plural)

    display_fields = modeladmin.get_list_display(request)
    headers = []
    for field in display_fields:
        method = getattr(modeladmin, field, None) if isinstance(field, str) else field
        label = getattr(method, 'short_description', None) or (
            opts.get_field(field).verbose_name if isinstance(field, str) else field
        )
        headers.append(str(label))
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='2E86AB')
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for obj in queryset:
        row = []
        for field in display_fields:
            if callable(field):
                value = field(obj)
            elif hasattr(modeladmin, field):
                value = getattr(modeladmin, field)(obj)
            else:
                value = getattr(obj, field, '')
            row.append(_strip_html(value) if value is not None else '')
        ws.append(row)

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(response)
    return response

export_to_excel.short_description = 'Экспортировать в Excel'

BONUS_PER_NORM = 300 
NORM_SIZE      = 30   

def export_bonuses(modeladmin, request, queryset):  
    """Отчёт по премиям: по каждому поисковику — разбивка по магазинам."""
    searchmen = queryset.filter(role='searchman')

    rows = (
        Product.objects
        .filter(creator__in=searchmen)
        .values(
            'creator__id', 'creator__name', 'creator__phone',
            'store__name',
        )
        .annotate(cnt=Count('id'))
        .order_by('creator__name', 'store__name')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Премии'

    # ── Стили ──
    blue_fill   = PatternFill('solid', fgColor='2E86AB')
    green_fill  = PatternFill('solid', fgColor='A8D5BA')
    yellow_fill = PatternFill('solid', fgColor='FFE066')
    gray_fill   = PatternFill('solid', fgColor='D9D9D9')
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _style(cell, fill=None, bold=False, align='left', color='000000'):
        cell.font = Font(bold=bold, color=color)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        cell.border = border
        if fill:
            cell.fill = fill

    headers = ['Сотрудник', 'Телефон', 'Магазин',
               'Загружено товаров', f'Норм (×{NORM_SIZE})', f'Премия (сом)']
    ws.append(headers)
    ws.row_dimensions[1].height = 22
    for col_idx, _ in enumerate(headers, 1):
        _style(ws.cell(1, col_idx), fill=blue_fill, bold=True, align='center', color='FFFFFF')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16

    current_staff_id = None
    staff_bonus = 0
    grand_total = 0
    data_rows = list(rows)

    def _flush_staff_total(row_num, name, bonus):
        ws.append(['', '', f'Итого — {name}', '', '', bonus])
        r = ws.max_row
        for c in range(1, 7):
            _style(ws.cell(r, c), fill=green_fill, bold=True, align='right' if c == 6 else 'left')
        ws.cell(r, 6).number_format = '#,##0'

    for i, row in enumerate(data_rows):
        sid    = row['creator__id']
        name   = row['creator__name']
        phone  = str(row['creator__phone'])
        store  = row['store__name']
        cnt    = row['cnt']
        norms  = cnt // NORM_SIZE
        bonus  = norms * BONUS_PER_NORM

        if current_staff_id is not None and sid != current_staff_id:
            _flush_staff_total(ws.max_row, prev_name, staff_bonus)
            staff_bonus = 0

        current_staff_id = sid
        prev_name = name
        staff_bonus += bonus
        grand_total += bonus

        show_name  = name  if (i == 0 or data_rows[i-1]['creator__id'] != sid) else ''
        show_phone = phone if (i == 0 or data_rows[i-1]['creator__id'] != sid) else ''

        ws.append([show_name, show_phone, store, cnt, norms, bonus])
        r = ws.max_row
        row_fill = yellow_fill if bonus > 0 else None
        for c in range(1, 7):
            _style(ws.cell(r, c), fill=row_fill)
        ws.cell(r, 4).alignment = Alignment(horizontal='center')
        ws.cell(r, 5).alignment = Alignment(horizontal='center')
        ws.cell(r, 6).number_format = '#,##0'

    if current_staff_id is not None:
        _flush_staff_total(ws.max_row, prev_name, staff_bonus)

    ws.append([])
    ws.append(['', '', 'ИТОГО ПРЕМИЙ', '', '', grand_total])
    r = ws.max_row
    for c in range(1, 7):
        _style(ws.cell(r, c), fill=gray_fill, bold=True, align='right' if c == 6 else 'left')
    ws.cell(r, 6).number_format = '#,##0'

    ws.append([])
    ws.append([f'* 1 норма = {NORM_SIZE} товаров = {BONUS_PER_NORM} сом'])
    ws.append(['* Жёлтые строки — магазины с начисленной премией'])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Премии.xlsx'
    wb.save(response)
    return response

export_bonuses.short_description = '💰 Экспорт премий'


def export_daily_stats(modeladmin, request, queryset):  # noqa: ARG001
    """Кто сколько сделал за ТЕКУЩИЙ день: найдено / сделано ИИ / загружено."""
    today = timezone.localdate()

    wb = Workbook()
    ws = wb.active
    ws.title = 'За день'

    blue_fill = PatternFill('solid', fgColor='2E86AB')
    gray_fill = PatternFill('solid', fgColor='D9D9D9')
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _style(cell, fill=None, bold=False, align='left', color='000000'):
        cell.font = Font(bold=bold, color=color)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        cell.border = border
        if fill:
            cell.fill = fill

    ws.append([f'Статистика за {today.strftime("%d.%m.%Y")}'])
    ws.merge_cells('A1:E1')
    _style(ws.cell(1, 1), bold=True, align='center')

    headers = ['Сотрудник', 'Роль', 'Найдено (товаров)', 'Сделано (товаров)', 'Загружено (товаров)']
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        _style(ws.cell(2, col_idx), fill=blue_fill, bold=True, align='center', color='FFFFFF')

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    total_found = total_made = total_uploaded = 0

    for staff in queryset.order_by('role', 'name'):
        # Найдено — товары, созданные поисковиком сегодня
        found = Product.objects.filter(creator=staff, created_at__date=today).count()
        # Сделано — РАЗНЫЕ товары, для которых ИИ-креатор сделал фото сегодня
        # (премия за товар, а не за количество фото: 3 фото или 2 = 1 товар)
        made = (
            Product.objects
            .filter(aiimage__creator=staff, aiimage__created_at__date=today)
            .distinct()
            .count()
        )
        # Загружено — товары, загруженные в панель сегодня
        uploaded = Product.objects.filter(uploader=staff, uploaded_at__date=today).count()

        # Пропускаем тех, кто сегодня ничего не делал
        if not (found or made or uploaded):
            continue

        total_found += found
        total_made += made
        total_uploaded += uploaded

        ws.append([staff.name, staff.get_role_display(), found, made, uploaded])
        r = ws.max_row
        for c in range(1, 6):
            _style(ws.cell(r, c), align='center' if c >= 3 else 'left')

    ws.append(['ИТОГО', '', total_found, total_made, total_uploaded])
    r = ws.max_row
    for c in range(1, 6):
        _style(ws.cell(r, c), fill=gray_fill, bold=True, align='center' if c >= 3 else 'left')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=stats_{today.isoformat()}.xlsx'
    wb.save(response)
    return response

export_daily_stats.short_description = '📊 Статистика за сегодня (найдено/сделано/загружено)'


class _StaffFkFilter(admin.SimpleListFilter):
    """База: список сотрудников, реально имеющих связанные записи, + фильтр по FK.

    Список формируется по фактическим данным (кто создавал товары / ИИ-фото),
    а не по роли — так фильтр всегда виден, когда есть что показывать, и
    отражает именно тех, у кого есть товары.
    """
    field_path = None    # путь фильтрации в queryset
    distinct = False     # True для обратных связей (один товар — много ИИ-фото)

    def staff_queryset(self):
        raise NotImplementedError

    def lookups(self, request, model_admin):
        return [(s.id, s.name) for s in self.staff_queryset().order_by('name')]

    def queryset(self, request, queryset):
        if self.value():
            qs = queryset.filter(**{self.field_path: self.value()})
            return qs.distinct() if self.distinct else qs
        return queryset


# ── Фильтры для ТОВАРОВ (количество Product) ──
class ProductSearchmanFilter(_StaffFkFilter):
    title = 'Поисковик'
    parameter_name = 'searchman'
    field_path = 'creator_id'

    def staff_queryset(self):
        return Staff.objects.filter(found_products__isnull=False).distinct()


class AiCreatorFilter(_StaffFkFilter):
    """Фильтр товаров по ИИ-креатору — считает именно товары, не фото."""
    title = 'ИИ-креатор'
    parameter_name = 'ai_creator'
    field_path = 'aiimage__creator_id'
    distinct = True

    def staff_queryset(self):
        return Staff.objects.filter(aiimage__isnull=False).distinct()


# ── Фильтры для ИИ-ИЗОБРАЖЕНИЙ ──
class AiImageAiCreatorFilter(_StaffFkFilter):
    title = 'ИИ-креатор'
    parameter_name = 'ai_creator'
    field_path = 'creator_id'

    def staff_queryset(self):
        return Staff.objects.filter(aiimage__isnull=False).distinct()


class AiImageSearchmanFilter(_StaffFkFilter):
    title = 'Поисковик'
    parameter_name = 'searchman'
    field_path = 'product__creator_id'

    def staff_queryset(self):
        return Staff.objects.filter(found_products__aiimage__isnull=False).distinct()


class ProductImageInline(admin.TabularInline):
    """Фото товара (M2M) — просмотр и удаление отдельных фото."""
    model = Product.images.through
    extra = 0
    fields = ('preview', 'image_path')
    readonly_fields = ('preview', 'image_path')
    verbose_name = 'Фото товара'
    verbose_name_plural = 'Фото товара'
    can_add = False

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('image')

    @admin.display(description='Превью')
    def preview(self, obj):
        return _safe_img(obj.image.image if obj.image else None, 70)

    @admin.display(description='Файл')
    def image_path(self, obj):
        if not obj.image:
            return '—'
        return format_html(
            '<a href="/admin/staff/image/{}/change/" target="_blank">{}</a>',
            obj.image_id, obj.image.image.name,
        )


class ProductInlineForStore(admin.TabularInline):
    model = Product
    fk_name = 'store'
    extra = 0
    fields = ('name', 'creator', 'main_image_preview', 'images_count',
              'size', 'color', 'created_at')
    readonly_fields = ('main_image_preview', 'images_count', 'created_at')
    show_change_link = True

    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'creator', 'main_image'
        ).prefetch_related('images')

    @admin.display(description='Главное фото')
    def main_image_preview(self, obj):
        return _safe_img(obj.main_image.image if obj.main_image else None, 45)

    @admin.display(description='Доп. фото')
    def images_count(self, obj):
        n = obj.images.count()
        return f'{n} шт.' if n else '—'


class ProductInlineForStaff(admin.TabularInline):
    model = Product
    fk_name = 'creator'
    extra = 0
    fields = ('name', 'store', 'main_image_preview', 'images_count', 'created_at')
    readonly_fields = ('main_image_preview', 'images_count', 'created_at')
    show_change_link = True
    can_add = False

    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'store', 'main_image'
        ).prefetch_related('images')

    @admin.display(description='Главное фото')
    def main_image_preview(self, obj):
        return _safe_img(obj.main_image.image if obj.main_image else None, 45)

    @admin.display(description='Доп. фото')
    def images_count(self, obj):
        n = obj.images.count()
        return f'{n} шт.' if n else '—'


class AiImageInlineForProduct(admin.TabularInline):
    model = AiImage
    fk_name = 'product'
    extra = 0
    fields = ('preview', 'creator', 'created_at')
    readonly_fields = ('preview', 'created_at')
    show_change_link = True

    @admin.display(description='Превью')
    def preview(self, obj):
        return _safe_img(obj.image, 45)


class AiImageInlineForStaff(admin.TabularInline):
    model = AiImage
    fk_name = 'creator'
    extra = 0
    fields = ('preview', 'product', 'created_at')
    readonly_fields = ('preview', 'created_at')
    show_change_link = True
    can_add = False

    @admin.display(description='Превью')
    def preview(self, obj):
        return _safe_img(obj.image, 45)


@admin.register(Store)
class StoreAdmin(admin.ModelAdmin):
    list_display = ('name', 'phone', 'product_count', 'created_at')
    list_display_links = ('name',)
    list_filter = (('created_at', admin.DateFieldListFilter),)
    search_fields = ('name', 'phone')
    inlines = [ProductInlineForStore]
    actions = [export_to_excel]
    list_per_page = 25
    date_hierarchy = 'created_at'
    readonly_fields = ('created_at', 'updated_at')

    fieldsets = (
        (None, {'fields': ('name', 'phone')}),
        ('Даты', {'classes': ('collapse',), 'fields': ('created_at', 'updated_at')}),
    )

    def get_queryset(self, request):
        return super().get_queryset(request).annotate(_cnt=Count('product'))

    @admin.display(description='Товаров', ordering='_cnt')
    def product_count(self, obj):
        return obj._cnt


@admin.register(Staff)
class StaffAdmin(admin.ModelAdmin):
    list_display = ('name', 'phone', 'role', 'lang', 'registred', 'tg_id', 'created_at')
    list_display_links = ('name',)
    list_editable = ('registred',)
    list_filter = ('role', 'lang', 'registred', ('created_at', admin.DateFieldListFilter))
    search_fields = ('name', 'phone', 'tg_id')
    inlines = [ProductInlineForStaff, AiImageInlineForStaff]
    actions = [export_daily_stats, export_bonuses, export_to_excel]
    list_per_page = 25
    date_hierarchy = 'created_at'
    readonly_fields = ('created_at', 'updated_at')

    fieldsets = (
        (None, {'fields': ('name', 'phone', 'role', 'lang')}),
        ('Дополнительно', {
            'classes': ('collapse',),
            'fields': ('age', 'tg_id', 'registred'),
        }),
        ('Даты', {'classes': ('collapse',), 'fields': ('created_at', 'updated_at')}),
    )


@admin.register(Image)
class ImageAdmin(admin.ModelAdmin):
    list_display = ('preview', 'image', 'created_at')
    list_display_links = ('preview', 'image')
    list_filter = (('created_at', admin.DateFieldListFilter),)
    search_fields = ('image',)
    readonly_fields = ('preview', 'created_at', 'updated_at')
    actions = [export_to_excel]
    list_per_page = 30

    fieldsets = (
        (None, {'fields': ('image', 'preview')}),
        ('Даты', {'classes': ('collapse',), 'fields': ('created_at', 'updated_at')}),
    )

    @admin.display(description='Превью')
    def preview(self, obj):
        return _safe_img(obj.image, 80)


@admin.register(AiImage)
class AiImageAdmin(admin.ModelAdmin):
    list_display = ('preview', 'creator_link', 'product_link', 'created_at')
    list_display_links = ('preview',)
    list_filter = (
        AiImageAiCreatorFilter,
        AiImageSearchmanFilter,
        ('created_at', admin.DateFieldListFilter),
    )
    search_fields = ('creator__name', 'creator__phone', 'product__name',
                     'product__creator__name')
    readonly_fields = ('preview', 'created_at', 'updated_at')
    raw_id_fields = ('creator', 'product')
    actions = [export_to_excel]
    list_per_page = 25
    date_hierarchy = 'created_at'

    fieldsets = (
        (None, {'fields': ('creator', 'product', 'image', 'preview')}),
        ('Даты', {'classes': ('collapse',), 'fields': ('created_at', 'updated_at')}),
    )

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('creator', 'product')

    @admin.display(description='Превью')
    def preview(self, obj):
        return _safe_img(obj.image, 80)

    @admin.display(description='Создатель')
    def creator_link(self, obj):
        return format_html(
            '<a href="/admin/staff/staff/{}/change/">{}</a>',
            obj.creator_id, obj.creator.name,
        )

    @admin.display(description='Товар')
    def product_link(self, obj):
        return format_html(
            '<a href="/admin/staff/product/{}/change/">{}</a>',
            obj.product_id, obj.product.name,
        )


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = (
        'name', 'price', 'store_link', 'creator_link', 'uploader_link',
        'main_image_preview', 'images_count', 'created_at', 'uploaded_at',
    )
    list_display_links = ('name',)
    list_filter = (
        ProductSearchmanFilter,
        AiCreatorFilter,
        'uploader',
        'store',
        ('created_at', admin.DateFieldListFilter),
        ('uploaded_at', admin.DateFieldListFilter),
    )
    search_fields = ('name', 'price', 'store__name', 'creator__name',
                     'uploader__name', 'characteristics')
    readonly_fields = ('main_image_preview', 'images_gallery', 'created_at', 'updated_at')
    inlines = [ProductImageInline, AiImageInlineForProduct]
    actions = [export_to_excel]
    save_on_top = True
    list_per_page = 25
    date_hierarchy = 'created_at'
    raw_id_fields = ('store', 'creator', 'main_image', 'uploader')

    fieldsets = (
        ('Основное', {'fields': ('name', 'price', 'store', 'creator')}),
        ('Загрузка в панель', {'fields': ('uploader', 'uploaded_at')}),
        ('Изображения', {
            'fields': ('main_image', 'main_image_preview', 'images_gallery'),
            'description': 'Доп. фото управляются ниже через секцию «Фото товара».',
        }),
        ('Характеристики', {
            'fields': ('size', 'color', 'material', 'characteristics', 'packaging'),
        }),
        ('Даты', {'classes': ('collapse',), 'fields': ('created_at', 'updated_at')}),
    )

    def get_readonly_fields(self, request, obj=None):
        return self.readonly_fields + ('uploaded_at',)

    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'store', 'creator', 'uploader', 'main_image',
        ).prefetch_related('images')

    @admin.display(description='Магазин')
    def store_link(self, obj):
        return format_html(
            '<a href="/admin/staff/store/{}/change/">{}</a>',
            obj.store_id, obj.store.name,
        )

    @admin.display(description='Поисковик')
    def creator_link(self, obj):
        return format_html(
            '<a href="/admin/staff/staff/{}/change/">{}</a>',
            obj.creator_id, obj.creator.name,
        )

    @admin.display(description='Загрузчик')
    def uploader_link(self, obj):
        if not obj.uploader_id:
            return '—'
        return format_html(
            '<a href="/admin/staff/staff/{}/change/">{}</a>',
            obj.uploader_id, obj.uploader.name,
        )

    @admin.display(description='Главное фото')
    def main_image_preview(self, obj):
        return _safe_img(obj.main_image.image if obj.main_image else None, 60)

    @admin.display(description='Доп. фото')
    def images_count(self, obj):
        n = obj.images.count()
        return f'{n} шт.' if n else '—'

    @admin.display(description='Галерея фото товара')
    def images_gallery(self, obj):
        parts = []
        for img in obj.images.all():
            try:
                parts.append(_img(img.image.url, 90))
            except (ValueError, AttributeError):
                pass
        return mark_safe(''.join(parts)) if parts else '—'
